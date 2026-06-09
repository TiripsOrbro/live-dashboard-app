#!/usr/bin/env python3
"""Export build-to workbook rows to config/build-to-workbook.json (no Excel at runtime)."""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "buildto-3811-copy.xlsx"
OUT = ROOT / "config" / "build-to-workbook.json"

SHEETS = {
    "DRY": {"orderCol": 1, "sohCol": 2, "perPackCol": 4, "nameCol": 5, "mmxCol": 12, "layout": "dry"},
    "FRIDGE & FREEZER": {"orderCol": 1, "sohCol": 2, "perPackCol": 3, "nameCol": 4, "mmxCol": 11, "layout": "fridge"},
    "SCHWEPPES": {"orderCol": 1, "sohCol": 2, "perPackCol": 3, "nameCol": 4, "mmxCol": 11, "layout": "fridge"},
    "BEGA": {"orderCol": 1, "sohCol": 2, "perPackCol": 3, "nameCol": 4, "mmxCol": 11, "layout": "fridge"},
    "CUTFRESH": {"orderCol": 1, "sohCol": 2, "perPackCol": 3, "nameCol": 4, "mmxCol": 11, "layout": "cutfresh"},
}

SKIP_NAMES = {"ITEM", "DRY STOCK BUILD TO GUIDE", "FRIDGE & FREEZER BUILD TO GUIDE"}


def norm_code(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.startswith("="):
        return ""
    digits = re.sub(r"\D", "", s)
    return digits or s


def infer_build_rule(build_f, need_f, name, layout, fixed_build_to):
    name_u = str(name or "").upper()
    build_f = str(build_f or "")
    need_f = str(need_f or "")

    if layout == "cutfresh":
        return {"type": "cutfresh"}

    if fixed_build_to is not None and isinstance(fixed_build_to, (int, float)):
        if "BIB" in name_u or "FROZEN" in name_u:
            return {"type": "fixed", "value": float(fixed_build_to)}

    if "PACKS" in name_u or "GLOVES" in name_u:
        return {"type": "pack10", "innerPerCarton": 10, "onOrderCartonFactor": 10}
    if "PAPER TOWEL" in name_u:
        return {"type": "pack10", "innerPerCarton": 6, "onOrderCartonFactor": 12}
    if "TOILET PAPER" in name_u:
        return {"type": "pack10", "innerPerCarton": 12, "onOrderCartonFactor": 6}

    if "+2" in build_f or "BIG BELL" in name_u:
        return {"type": "days10add2", "add": 2}

    if need_f and ">10" not in need_f and "IF(H" in need_f:
        return {"type": "days10", "skipDaysHoldingCap": True}

    return {"type": "days10"}


def export_cutfresh_calendar(ws_val):
    """Export weekday → item → build-to from the CUTFRESH delivery matrix."""
    cal = {}
    # Row 21-24: Monday / Wednesday / Saturday block (cols D=3, G=6, J=9)
    blocks = [
        ("MONDAY", 21, 3, 1),
        ("WEDNESDAY", 21, 6, 1),
        ("SATURDAY", 21, 9, 1),
        ("TUESDAY", 27, 3, 1),
        ("THURSDAY", 27, 6, 1),
        ("SUNDAY", 27, 9, 2),  # Sunday value column offset differs in sheet
    ]
    for day, start_row, col, val_offset in blocks:
        cal[day] = {}
        for offset in range(4):
            row = list(ws_val.iter_rows(min_row=start_row + offset, max_row=start_row + offset, values_only=True))[0]
            item = str(row[col] or "").strip().upper().rstrip()
            val = row[col + val_offset] if len(row) > col + val_offset else None
            if item and isinstance(val, (int, float)):
                cal[day][item] = float(val)
    return cal


def export_rows(xlsx_path):
    wb_f = load_workbook(xlsx_path, data_only=False, read_only=True)
    wb_v = load_workbook(xlsx_path, data_only=True, read_only=True)
    rows = []
    cutfresh_calendar = None

    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb_f.sheetnames:
            continue
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]
        layout = cfg["layout"]

        if layout == "cutfresh":
            cutfresh_calendar = export_cutfresh_calendar(ws_v)

        for i, row_cells in enumerate(ws_f.iter_rows(min_row=5, max_row=200, values_only=False), start=5):
            vals_f = [c.value for c in row_cells]
            vals_v = list(ws_v.iter_rows(min_row=i, max_row=i, values_only=True))[0]
            if not any(v is not None and str(v).strip() for v in vals_f):
                continue
            name = vals_f[cfg["nameCol"]]
            if not name or str(name).startswith("="):
                continue
            name_s = str(name).strip()
            if name_s.upper() in SKIP_NAMES:
                continue
            if "openpyxl" in name_s or name_s.lower() in ("monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"):
                continue
            if layout == "cutfresh" and name_s.upper() not in ("LETTUCE", "ONION", "CORIANDER", "TOMATO", "TOMATO "):
                continue

            order_code = norm_code(vals_f[cfg["orderCol"]])
            soh_label = str(vals_f[cfg["sohCol"]] or "").strip()
            if soh_label in ("#N/A", "N/A"):
                soh_label = ""
            per_pack = vals_f[cfg["perPackCol"]]
            mmx = norm_code(vals_f[cfg["mmxCol"]])
            if isinstance(per_pack, str) and per_pack.startswith("="):
                per_pack = None
            try:
                per_pack_n = float(per_pack) if per_pack is not None else None
            except (TypeError, ValueError):
                per_pack_n = None

            if layout == "dry":
                daily_c, bt_c, need_c = 6, 10, 11
            else:
                daily_c, bt_c, need_c = 5, 9, 10

            daily_cell = vals_f[daily_c] if len(vals_f) > daily_c else None
            build_cell = vals_f[bt_c] if len(vals_f) > bt_c else None
            need_cell = vals_f[need_c] if len(vals_f) > need_c else None
            fixed_build_to = vals_v[bt_c] if len(vals_v) > bt_c else None
            if isinstance(fixed_build_to, str):
                fixed_build_to = None

            daily_manual = None
            if isinstance(daily_cell, (int, float)):
                daily_manual = float(daily_cell)
            elif isinstance(daily_cell, str) and not daily_cell.startswith("="):
                try:
                    daily_manual = float(daily_cell)
                except ValueError:
                    pass

            rule = infer_build_rule(build_cell, need_cell, name_s, layout, fixed_build_to)
            if rule["type"] == "fixed":
                rule["value"] = float(fixed_build_to)

            rows.append({
                "sheet": sheet_name,
                "name": name_s,
                "orderCode": order_code if order_code != "###" else "",
                "sohLabel": soh_label,
                "mmxCode": mmx,
                "perPack": per_pack_n,
                "dailyManual": daily_manual,
                "buildToRule": rule,
            })

    wb_f.close()
    wb_v.close()
    return rows, cutfresh_calendar


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"Missing {xlsx}", file=sys.stderr)
        sys.exit(1)
    rows, cutfresh_calendar = export_rows(xlsx)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "source": xlsx.name,
        "rows": rows,
    }
    if cutfresh_calendar:
        payload["cutfreshCalendar"] = cutfresh_calendar
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} rows to {OUT}")


if __name__ == "__main__":
    main()
