#!/usr/bin/env python3
"""Dump Excel formulas from build-to workbook (not cached values)."""
import sys
from openpyxl import load_workbook

path = sys.argv[1] if len(sys.argv) > 1 else r"data/buildto-3811-copy.xlsx"
wb = load_workbook(path, data_only=False, read_only=True)

for sheet in ["DRY", "FRIDGE & FREEZER", "SCHWEPPES", "BEGA", "CUTFRESH"]:
    if sheet not in wb.sheetnames:
        continue
    ws = wb[sheet]
    print(f"\n========== {sheet} (sample formulas) ==========")
    count = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=120, values_only=False)):
        vals = [c.value for c in row]
        if not any(v is not None and str(v).strip() != "" for v in vals):
            continue
        # show rows with formulas in usage/buildto/need cols (F-K / 6-11)
        formulas = []
        for j, c in enumerate(row):
            if isinstance(c.value, str) and c.value.startswith("="):
                formulas.append((j, c.value))
        if formulas and count < 25:
            label = vals[5] or vals[4] or vals[1] or vals[0]
            print(f"Row {i+1} {label}:")
            for j, f in formulas:
                print(f"  col {j}: {f[:200]}")
            count += 1
wb.close()
